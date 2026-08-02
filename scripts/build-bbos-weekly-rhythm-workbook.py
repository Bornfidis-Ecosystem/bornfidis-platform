#!/usr/bin/env python3
"""Build BBOS Weekly Operating Rhythm Workbook V1.

Phase One companion workbook spanning Modules 1, 2, and 4.
References (does not duplicate) the Pricing & Margin Calculator V1.0.

Output: content/bbos/bbos-weekly-operating-rhythm-workbook-v1.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "content" / "bbos" / "bbos-weekly-operating-rhythm-workbook-v1.xlsx"

ASSET = "BBOS Weekly Operating Rhythm Workbook"
VERSION = "Version 1.0"
CYCLE = "Plan → Deliver → Sell and Follow Up → Review → Close"

INK = "1C1C1C"
FOREST = "1A3C34"
GOLD = "C9A84C"
PAPER = "F5F1E6"
WHITE = "FFFFFF"
LINE = "E5E3DC"
MUTED = "6B7280"

thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
gold_border = Border(
    left=Side(style="medium", color=GOLD),
    right=Side(style="medium", color=GOLD),
    top=Side(style="medium", color=GOLD),
    bottom=Side(style="medium", color=GOLD),
)

fill_forest = PatternFill("solid", fgColor=FOREST)
fill_paper = PatternFill("solid", fgColor=PAPER)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_gold_light = PatternFill("solid", fgColor="F8F1D8")

font_title = Font(name="Calibri", size=16, bold=True, color=INK)
font_h = Font(name="Calibri", size=12, bold=True, color=WHITE)
font_section = Font(name="Calibri", size=12, bold=True, color=FOREST)
font_label = Font(name="Calibri", size=11, bold=True, color=INK)
font_body = Font(name="Calibri", size=11, color=INK)
font_muted = Font(name="Calibri", size=10, italic=True, color=MUTED)
font_gold = Font(name="Calibri", size=10, bold=True, color=FOREST)
font_footer = Font(name="Calibri", size=9, color=MUTED)

prot_lock = Protection(locked=True)
prot_unlock = Protection(locked=False)


def set_widths(ws, widths: dict[str, float]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def header_block(ws, title: str, supports: str, instruction: str, cols: int = 6):
    ws["A1"] = "BBOS"
    ws["A1"].font = Font(name="Calibri", size=10, bold=True, color=FOREST)
    ws["B1"] = ASSET
    ws["B1"].font = font_gold
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=min(cols, 4))

    ws["A2"] = title
    ws["A2"].font = font_title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)

    ws["A3"] = supports
    ws["A3"].font = font_gold

    ws["A4"] = instruction
    ws["A4"].font = font_muted
    ws["A4"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=cols)
    ws.row_dimensions[4].height = 36

    ws["E1"] = "Date / Week #"
    ws["E1"].font = font_label
    input_cell(ws, 1, 6, "")


def footer(ws, row: int, supports: str, cols: int = 6):
    ws.cell(row, 1, f"{ASSET} · {VERSION} · {supports} · Not legal, tax, accounting, or financial advice.")
    ws.cell(row, 1).font = font_footer
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)


def input_cell(ws, row: int, col: int, value=None):
    cell = ws.cell(row, col, value if value is not None else "")
    cell.fill = fill_white
    cell.border = gold_border
    cell.font = font_body
    cell.protection = prot_unlock
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def label(ws, row: int, col: int, text: str):
    cell = ws.cell(row, col, text)
    cell.font = font_label
    cell.protection = prot_lock
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def note(ws, row: int, col: int, text: str, cols: int = 1):
    cell = ws.cell(row, col, text)
    cell.font = font_muted
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.protection = prot_lock
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + cols - 1)
    return cell


def section(ws, row: int, text: str, cols: int = 6):
    cell = ws.cell(row, 1, text)
    cell.font = font_section
    cell.fill = fill_paper
    cell.protection = prot_lock
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    return row + 1


def style_header_row(ws, row: int, headers: list[str]):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row, i, h)
        cell.fill = fill_forest
        cell.font = font_h
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.protection = prot_lock
    ws.row_dimensions[row].height = 28


def blank_input_rows(ws, start: int, n: int, cols: int, height: int = 22):
    for r in range(start, start + n):
        for c in range(1, cols + 1):
            input_cell(ws, r, c, "")
        ws.row_dimensions[r].height = height


def tall_field(ws, row: int, label_text: str, height: int = 48, cols: int = 6):
    label(ws, row, 1, label_text)
    input_cell(ws, row, 2, "")
    if cols > 2:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = height


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------


def build_start_here(wb: Workbook):
    ws = wb.create_sheet("Start Here", 0)
    set_widths(ws, {"A": 92})
    ws["A1"] = f"{ASSET} — Plan, Run, Close Your Week"
    ws["A1"].font = font_title
    lines = [
        "",
        VERSION,
        "",
        "Purpose",
        "Hold the practical outputs of Modules 1, 2, and 4 in one workbook so you can",
        "run your week without rebuilding tools every Monday.",
        "",
        f"BBOS Week cycle: {CYCLE}",
        "",
        "What this workbook includes",
        "Module 1 — Cost of Disorder · Owner Time Map · Business Snapshot",
        "Module 2 — Offer Builder · Offer Ladder · Offer One-Pager · Pipeline ·",
        "           Should I Say Yes? · Follow-Up Cadence",
        "Module 4 — My Rhythm · Week Plan · Daily Anchors · Delivery Quality ·",
        "           Pricing Checkpoint · Weekly Close (with Next-Week Handoff) ·",
        "           Catch-Up Protocol · 13-Week Tracker",
        "",
        "What this workbook does NOT include",
        "Pricing math lives in the separate Pricing & Margin Calculator V1.0",
        "(Module 3). Open that file when you need a price. Do not rebuild Cost Stack,",
        "Target Margin, or Guardrail math here.",
        "",
        "Suggested order (first pass)",
        "1. Cost of Disorder → Owner Time Map → Business Snapshot",
        "2. Offer Builder → Offer Ladder → Offer One-Pager",
        "3. Pipeline Tracker → Should I Say Yes → Follow-Up Cadence",
        "4. My Rhythm → Daily Anchors → Delivery Quality → Catch-Up Protocol",
        "5. Each week: Week Plan → run delivery/Sell → Pricing Checkpoint → Weekly Close",
        "6. Mark progress on 13-Week Tracker",
        "",
        "Duration governance (Module 4)",
        "Weekly Planning Block: 25–40 minutes (target ~30)",
        "Pipeline and Follow-Up Block: 30–60 minutes (target ~45)",
        "Weekly Close: 25–45 minutes (hard ceiling 45 in a normal week;",
        "  stretch to 60 only in catch-up or consolidation weeks)",
        "Daily Anchors: 2–3 only",
        "",
        "Financial boundary",
        "Module 4 notices; Module 5 analyzes.",
        "Pricing Checkpoint and Close money lines notice only — no runway,",
        "cash-flow planning, or Core Number Set analysis in this workbook.",
        "",
        "Privacy",
        "Do not put client passwords, access codes, private financial details,",
        "or real third-party confidential data in this file.",
        "",
        "Example — Maya sheet is fictional illustration only.",
    ]
    for i, text in enumerate(lines, start=2):
        cell = ws.cell(i, 1, text)
        if text in {
            "Purpose",
            "What this workbook includes",
            "What this workbook does NOT include",
            "Suggested order (first pass)",
            "Duration governance (Module 4)",
            "Financial boundary",
            "Privacy",
        }:
            cell.font = font_section
        elif text == VERSION:
            cell.font = font_gold
        else:
            cell.font = font_body
    footer(ws, len(lines) + 3, "Supports Modules 1, 2, and 4", 1)


def build_cost_of_disorder(wb: Workbook):
    ws = wb.create_sheet("Cost of Disorder")
    set_widths(ws, {"A": 42, "B": 14, "C": 22, "D": 28, "E": 12, "F": 12})
    header_block(
        ws,
        "Cost-of-Disorder Audit",
        "Supports Module 1 — Business Foundation",
        "List 6–9 concrete leaks from the last 30 days across Time, Money, and Capacity. "
        "Do not sum currencies. Name the single most expensive pattern at the bottom.",
    )
    r = section(ws, 6, "Audit entries (aim 6–9)", 4)
    style_header_row(ws, r, ["Leak (specific)", "Currency", "Rough estimate", "Only I can handle it?"])
    blank_input_rows(ws, r + 1, 9, 4, 28)
    dv = DataValidation(type="list", formula1='"Time,Money,Capacity"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"B{r+1}:B{r+9}")
    dv2 = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add(f"D{r+1}:D{r+9}")
    end = r + 11
    note(ws, end, 1, "Currencies stay separate. Hours, money, and lost jobs are not one total.", 4)
    end += 2
    label(ws, end, 1, "The single most expensive pattern here is:")
    ws.merge_cells(start_row=end, start_column=2, end_row=end, end_column=4)
    input_cell(ws, end, 2, "")
    ws.row_dimensions[end].height = 40
    footer(ws, end + 2, "Supports Module 1", 4)


def build_owner_time_map(wb: Workbook):
    ws = wb.create_sheet("Owner Time Map")
    set_widths(ws, {"A": 14, "B": 12, "C": 12, "D": 12, "E": 12, "F": 28})
    header_block(
        ws,
        "Owner's Time Map",
        "Supports Module 1 — Business Foundation",
        "Log rough hours for one normal week into four buckets: Doing, Selling, Running, Building. "
        "Minimum three honest days; ideally seven.",
    )
    r = 6
    style_header_row(ws, r, ["Day", "Doing", "Selling", "Running", "Building", "Notes"])
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for i, day in enumerate(days):
        row = r + 1 + i
        label(ws, row, 1, day)
        for c in range(2, 6):
            input_cell(ws, row, c, "")
        input_cell(ws, row, 6, "")
    tot = r + 9
    label(ws, tot, 1, "Week totals")
    for c, col in enumerate(["B", "C", "D", "E"], start=2):
        cell = ws.cell(tot, c, f"=SUM({col}{r+1}:{col}{r+7})")
        cell.fill = fill_paper
        cell.border = thin
        cell.font = font_label
        cell.protection = prot_lock
    end = tot + 2
    tall_field(ws, end, "The bucket that surprised me most was ______, because ______.", 48, 6)
    footer(ws, end + 2, "Supports Module 1", 6)


def build_snapshot(wb: Workbook):
    ws = wb.create_sheet("Business Snapshot")
    set_widths(ws, {"A": 36, "B": 70, "C": 12, "D": 12, "E": 12, "F": 12})
    header_block(
        ws,
        "One-Page Business Snapshot",
        "Supports Module 1 — Business Foundation",
        "Fill all six fields. Keep each short. No client names, passwords, or private customer data. "
        "Choose one first weakness — not five.",
    )
    fields = [
        "1. The business I am really in",
        "2. My ideal customer — and who is a poor fit",
        "3. My core offer — what's included + the outcome",
        "4. How I run today (honest)",
        "5. My biggest costs of disorder (top 2–3)",
        "6. My first weakness to fix (ONE only)",
    ]
    row = 6
    for f in fields:
        tall_field(ws, row, f, 52, 6)
        row += 1
    note(ws, row, 1, "Private working document. Revisit weekly in your Planning Block.", 6)
    footer(ws, row + 2, "Supports Module 1", 6)


def build_offer_builder(wb: Workbook):
    ws = wb.create_sheet("Offer Builder")
    set_widths(ws, {"A": 28, "B": 72, "C": 12, "D": 12, "E": 12, "F": 12})
    header_block(
        ws,
        "Offer Builder",
        "Supports Module 2 — Offer and Sales System",
        "Package your core offer. Name Optional items but do not price them here. "
        "Prices belong in Module 3 / Pricing & Margin Calculator.",
    )
    fields = [
        "Who it is for",
        "Intended outcome",
        "Included",
        "Optional (named, unpriced)",
        "Not Included",
    ]
    row = 6
    for f in fields:
        tall_field(ws, row, f, 56, 6)
        row += 1
    footer(ws, row + 1, "Supports Module 2", 6)


def build_offer_ladder(wb: Workbook):
    ws = wb.create_sheet("Offer Ladder")
    set_widths(ws, {"A": 14, "B": 36, "C": 36, "D": 28, "E": 12, "F": 12})
    header_block(
        ws,
        "Offer Ladder (3 rungs)",
        "Supports Module 2 — Offer and Sales System",
        "Exactly three rungs around one core offer. Use price placeholder until Module 3. "
        "Do not invent calculator math on this sheet.",
    )
    r = 6
    style_header_row(ws, r, ["Rung", "What it is", "Who it suits", "Price placeholder"])
    for i, rung in enumerate(["Entry", "Core", "Premium"]):
        row = r + 1 + i
        label(ws, row, 1, rung)
        for c in range(2, 4):
            input_cell(ws, row, c, "")
        input_cell(ws, row, 4, "[price — set in Module 3]")
        ws.row_dimensions[row].height = 48
    note(ws, r + 5, 1, "When ready to price, open Pricing & Margin Calculator V1.0 — then return here only to record the chosen quote.", 4)
    footer(ws, r + 7, "Supports Module 2", 4)


def build_one_pager(wb: Workbook):
    ws = wb.create_sheet("Offer One-Pager")
    set_widths(ws, {"A": 28, "B": 72, "C": 12, "D": 12, "E": 12, "F": 12})
    header_block(
        ws,
        "Offer One-Pager (draft content)",
        "Supports Module 2 — Offer and Sales System",
        "Draft the words for a one-page customer page. Keep it under a minute to read. "
        "No prices on this draft. One clear how-to-start action.",
    )
    fields = [
        "1. Headline",
        "2. Who it is for",
        "3. What you get (Included)",
        "4. Optional add-ons (unpriced)",
        "5. The ladder (Entry / Core / Premium in plain words)",
        "6. How to start (ONE action)",
    ]
    row = 6
    for f in fields:
        tall_field(ws, row, f, 48, 6)
        row += 1
    footer(ws, row + 1, "Supports Module 2", 6)


def build_pipeline(wb: Workbook):
    ws = wb.create_sheet("Pipeline Tracker")
    set_widths(ws, {"A": 26, "B": 12, "C": 12, "D": 28, "E": 28, "F": 18})
    header_block(
        ws,
        "Pipeline Tracker",
        "Supports Module 2 (setup) · Module 4 (weekly run)",
        "Stages: Lead → Qualified → Proposed → Won / Lost. Every live opportunity needs a next action. "
        "Not now stays in Lead with a revisit date. Not a revenue forecast — do not total pipeline worth.",
    )
    r = 6
    style_header_row(
        ws,
        r,
        ["Opportunity", "Stage", "Decision", "One-line reason", "Next action", "Next follow-up / revisit"],
    )
    blank_input_rows(ws, r + 1, 12, 6, 26)
    dv_stage = DataValidation(
        type="list", formula1='"Lead,Qualified,Proposed,Won,Lost"', allow_blank=True
    )
    ws.add_data_validation(dv_stage)
    dv_stage.add(f"B{r+1}:B{r+12}")
    dv_dec = DataValidation(type="list", formula1='"Yes,Not now,No"', allow_blank=True)
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"C{r+1}:C{r+12}")
    note(
        ws,
        r + 14,
        1,
        "Active opportunities only. Past-customer retention / referral follow-up belongs in Module 7.",
        6,
    )
    footer(ws, r + 16, "Supports Modules 2 and 4", 6)


def build_should_i_say_yes(wb: Workbook):
    ws = wb.create_sheet("Should I Say Yes")
    set_widths(ws, {"A": 22, "B": 14, "C": 40, "D": 16, "E": 18, "F": 18})
    header_block(
        ws,
        "Should I Say Yes? (Lead → Qualified gate)",
        "Supports Module 2 — Offer and Sales System",
        "Run five checks, then record Yes / Not now / No with one-line reason. "
        "Margin-viable is yes/no feasibility only — open the Pricing Calculator if you need numbers.",
    )
    r = section(ws, 6, "Five checks (per opportunity)", 5)
    style_header_row(ws, r, ["Opportunity", "Fit", "Boundary", "Capacity", "Margin-viable", "Red flags?"])
    # Need 6 cols for checks - adjust
    set_widths(ws, {"A": 22, "B": 10, "C": 12, "D": 12, "E": 14, "F": 12})
    blank_input_rows(ws, r + 1, 8, 6, 24)
    for col in ("B", "C", "D", "E"):
        dv = DataValidation(type="list", formula1='"Pass,Fail,Unsure"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}{r+1}:{col}{r+8}")
    dv_rf = DataValidation(type="list", formula1='"None,Some,Several"', allow_blank=True)
    ws.add_data_validation(dv_rf)
    dv_rf.add(f"F{r+1}:F{r+8}")

    end = r + 10
    end = section(ws, end, "Decision log", 5)
    style_header_row(ws, end, ["Opportunity", "Decision", "One-line reason", "Revisit date (if Not now)", "Notes"])
    # remap columns A-E for decision log
    for i in range(8):
        row = end + 1 + i
        for c in range(1, 6):
            input_cell(ws, row, c, "")
        ws.row_dimensions[row].height = 24
    dv_d = DataValidation(type="list", formula1='"Yes,Not now,No"', allow_blank=True)
    ws.add_data_validation(dv_d)
    dv_d.add(f"B{end+1}:B{end+8}")
    note(
        ws,
        end + 10,
        1,
        "Yes → Qualified. Not now → stay in Lead with revisit date. No → do not pursue.",
        5,
    )
    footer(ws, end + 12, "Supports Module 2", 5)


def build_follow_up_cadence(wb: Workbook):
    ws = wb.create_sheet("Follow-Up Cadence")
    set_widths(ws, {"A": 28, "B": 36, "C": 22, "D": 16, "E": 16, "F": 12})
    header_block(
        ws,
        "Follow-Up Cadence",
        "Supports Module 2 (write) · Module 4 (run weekly)",
        "Write 3–4 rules for active leads/opportunities only. Must cover new lead, sent proposal, "
        "final nudge, and a clear stop. Retention of past customers is Module 7.",
    )
    r = 6
    style_header_row(ws, r, ["Trigger", "Action", "Timing", "Pipeline stage (optional)"])
    blank_input_rows(ws, r + 1, 6, 4, 32)
    dv = DataValidation(
        type="list", formula1='"Lead,Qualified,Proposed,Any active"', allow_blank=True
    )
    ws.add_data_validation(dv)
    dv.add(f"D{r+1}:D{r+6}")
    end = r + 8
    tall_field(ws, end, "Stopping rule (when an opportunity moves to Lost)", 40, 4)
    footer(ws, end + 2, "Supports Modules 2 and 4", 4)


def build_my_rhythm(wb: Workbook):
    ws = wb.create_sheet("My Rhythm")
    set_widths(ws, {"A": 32, "B": 22, "C": 18, "D": 40, "E": 12, "F": 12})
    header_block(
        ws,
        "My Weekly Operating Rhythm",
        "Supports Module 4 — Weekly Operating Rhythm",
        "Schedule blocks that fit your delivery reality. No universal day or clock time. "
        "Write 4–6 Rhythm Protection Rules at the bottom.",
    )
    r = section(ws, 6, "Protected windows", 4)
    style_header_row(ws, r, ["Block", "Day / window", "Duration", "Notes"])
    blocks = [
        ("Weekly Planning Block", "25–40 min (target ~30)"),
        ("Pipeline and Follow-Up Block", "30–60 min (target ~45)"),
        ("Weekly Close", "25–45 min (ceiling 45)"),
    ]
    for i, (name, dur) in enumerate(blocks):
        row = r + 1 + i
        label(ws, row, 1, name)
        input_cell(ws, row, 2, "")
        input_cell(ws, row, 3, dur)
        input_cell(ws, row, 4, "")
        ws.row_dimensions[row].height = 28
    end = r + 5
    tall_field(ws, end, "These windows fit my delivery reality because:", 40, 4)
    end = section(ws, end + 2, "Rhythm Protection Rules (4–6)", 4)
    for i in range(6):
        label(ws, end + i, 1, f"Rule {i + 1}")
        ws.merge_cells(start_row=end + i, start_column=2, end_row=end + i, end_column=4)
        input_cell(ws, end + i, 2, "")
        ws.row_dimensions[end + i].height = 28
    note(
        ws,
        end + 7,
        1,
        "Starter ideas: protect Planning+Close; Sell may move but not vanish two weeks; anchors stay 2–3; "
        "do not negotiate price when exhausted; consolidation weeks stay lighter.",
        4,
    )
    footer(ws, end + 9, "Supports Module 4", 4)


def build_week_plan(wb: Workbook):
    ws = wb.create_sheet("Week Plan")
    set_widths(ws, {"A": 28, "B": 72, "C": 12, "D": 12, "E": 12, "F": 12})
    header_block(
        ws,
        "Weekly Planning Block",
        "Supports Module 4 — Weekly Operating Rhythm",
        "Duration 25–40 minutes (target ~30). Produce: top 3 outcomes, confirmed Sell+Close windows, "
        "one delivery risk. Do not rebuild offers or redesign prices here.",
    )
    row = 6
    tall_field(ws, row, "Snapshot glance (1 min) — anything changed?", 36, 6)
    row += 1
    tall_field(ws, row, "Capacity check (booked jobs / help / rest)", 40, 6)
    row += 1
    row = section(ws, row, "Top 3 outcomes (not 12)", 6)
    for i in range(1, 4):
        label(ws, row, 1, f"Outcome {i}")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        input_cell(ws, row, 2, "")
        ws.row_dimensions[row].height = 28
        row += 1
    tall_field(ws, row, "Confirm Pipeline/Follow-Up window", 28, 6)
    row += 1
    tall_field(ws, row, "Confirm Weekly Close window", 28, 6)
    row += 1
    tall_field(ws, row, "One delivery risk this week", 36, 6)
    row += 1
    note(ws, row, 1, "When the timer ends: write top 3 + windows + risk, then stop.", 6)
    footer(ws, row + 2, "Supports Module 4", 6)


def build_daily_anchors(wb: Workbook):
    ws = wb.create_sheet("Daily Anchors")
    set_widths(ws, {"A": 14, "B": 40, "C": 12, "D": 12, "E": 12, "F": 12})
    header_block(
        ws,
        "Daily Anchors (cap 2–3)",
        "Supports Module 4 — Weekly Operating Rhythm",
        "Choose 2–3 short, visible actions for most working days. Minutes, not hours. "
        "Miss one day → resume next. Do not build a Saturday guilt list.",
    )
    r = 6
    style_header_row(ws, r, ["#", "Daily Anchor (short action)", "Mon", "Tue", "Wed", "Thu"])
    # extend with Fri - need more columns
    set_widths(ws, {"A": 8, "B": 42, "C": 8, "D": 8, "E": 8, "F": 8, "G": 8, "H": 8})
    ws.cell(r, 7, "Fri").fill = fill_forest
    ws.cell(r, 7).font = font_h
    ws.cell(r, 8, "Sat/Sun").fill = fill_forest
    ws.cell(r, 8).font = font_h
    for i in range(1, 4):
        row = r + i
        label(ws, row, 1, str(i))
        input_cell(ws, row, 2, "")
        for c in range(3, 9):
            input_cell(ws, row, c, "")
        ws.row_dimensions[row].height = 28
    note(ws, r + 5, 1, "Leave Anchor 3 blank if two is enough. Cap remains 2–3.", 8)
    note(
        ws,
        r + 6,
        1,
        "Tick cells with ✓ on days completed. Anchors are not your whole day — delivery still dominates.",
        8,
    )
    footer(ws, r + 8, "Supports Module 4", 8)


def build_delivery_quality(wb: Workbook):
    ws = wb.create_sheet("Delivery Quality")
    set_widths(ws, {"A": 18, "B": 50, "C": 14, "D": 30, "E": 12, "F": 12})
    header_block(
        ws,
        "Delivery Quality Checklist",
        "Supports Module 4 — Weekly Operating Rhythm",
        "Routine around delivery — not full SOPs (Module 6). Write 3–5 quality bullets. "
        "Use Before / During / After.",
    )
    row = 6
    tall_field(ws, row, "Before — scope / Offer Ladder rung / Included standard", 40, 6)
    row += 1
    tall_field(ws, row, "During — deliver to that standard (notes)", 40, 6)
    row += 1
    row = section(ws, row, "After — quality ticks (3–5 bullets max)", 4)
    style_header_row(ws, row, ["#", "Quality check", "Met?", "Note / tell worker"])
    for i in range(1, 6):
        rr = row + i
        label(ws, rr, 1, str(i))
        input_cell(ws, rr, 2, "")
        input_cell(ws, rr, 3, "")
        input_cell(ws, rr, 4, "")
    dv = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C{row+1}:C{row+5}")
    footer(ws, row + 7, "Supports Module 4", 4)


def build_pricing_checkpoint(wb: Workbook):
    ws = wb.create_sheet("Pricing Checkpoint")
    set_widths(ws, {"A": 48, "B": 14, "C": 50, "D": 12, "E": 12, "F": 12})
    header_block(
        ws,
        "Money and Pricing Checkpoint (notice only)",
        "Supports Module 4 — Weekly Operating Rhythm",
        "Run inside the Pipeline and Follow-Up Block (~5–10 min). Module 4 notices; Module 5 analyzes. "
        "For price recalculation, open Pricing & Margin Calculator V1.0 — then return.",
    )
    r = 6
    style_header_row(ws, r, ["Notice question", "Status", "One-line note (not analysis)"])
    questions = [
        "Revenue recorded for jobs completed this week? (or gap noted)",
        "Any pricing Exception this week? Written with end date?",
        "Receipts or basic records missing?",
        "Cash question to park for Module 5 (not solve now)?",
    ]
    for i, q in enumerate(questions):
        row = r + 1 + i
        label(ws, row, 1, q)
        input_cell(ws, row, 2, "")
        input_cell(ws, row, 3, "")
        ws.row_dimensions[row].height = 36
    dv = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"B{r+1}:B{r+4}")
    note(
        ws,
        r + 6,
        1,
        "Do NOT use this sheet for: actual-margin analysis, cash-flow planning, runway, "
        "full Core Number Set, or redesigning Target Margin / Guardrails.",
        3,
    )
    note(
        ws,
        r + 7,
        1,
        "Reference tool: Pricing & Margin Calculator V1.0 (content/bbos/bbos-pricing-margin-calculator-v1.xlsx).",
        3,
    )
    footer(ws, r + 9, "Supports Module 4 (notice) · Module 5 analyzes later", 3)


def build_weekly_close(wb: Workbook):
    ws = wb.create_sheet("Weekly Close")
    set_widths(ws, {"A": 28, "B": 72, "C": 12, "D": 12, "E": 12, "F": 12})
    header_block(
        ws,
        "Weekly Close + Next-Week Handoff",
        "Supports Module 4 — Weekly Operating Rhythm",
        "Normal week: 25–45 minutes (hard ceiling 45). Catch-up/consolidation only: up to 60. "
        "Finish with a brief four-line Handoff — do not run full Planning here.",
    )
    agenda = [
        ("1. What shipped? (~5 min)", 40),
        ("2. Pipeline truth (~8–10 min)", 40),
        ("3. Follow-ups (~5 min)", 36),
        ("4. Money and pricing notice (~5 min) — notice only", 36),
        ("5. One lesson (one sentence, ~3 min)", 28),
    ]
    row = 6
    for text, h in agenda:
        tall_field(ws, row, text, h, 6)
        row += 1
    row = section(ws, row, "6. Next-Week Handoff (four lines only · ~5 min)", 6)
    handoff = [
        "H1. Next week’s top priority (one outcome)",
        "H2. First unresolved follow-up",
        "H3. First delivery risk or capacity concern",
        "H4. Date/window of next Weekly Planning Block",
    ]
    for text in handoff:
        tall_field(ws, row, text, 32, 6)
        row += 1
    label(ws, row, 1, "Time finished (stay ≤ 45 normal week)")
    input_cell(ws, row, 2, "")
    note(
        ws,
        row + 1,
        1,
        "If Handoff grows past four short lines, you are Planning early — move extras to Week Plan.",
        6,
    )
    footer(ws, row + 3, "Supports Module 4", 6)


def build_catch_up(wb: Workbook):
    ws = wb.create_sheet("Catch-Up Protocol")
    set_widths(ws, {"A": 36, "B": 64, "C": 12, "D": 12, "E": 12, "F": 12})
    header_block(
        ws,
        "Catch-Up Protocol",
        "Supports Module 4 — Weekly Operating Rhythm",
        "Write your protocol in 5–8 lines. Disruption is normal. Silence afterward creates drift. "
        "Catch-Up is professionalism — not a confession.",
    )
    row = section(ws, 6, "Minimum Catch-Up (15–25 min mid-disruption or same day)", 6)
    steps = [
        "1. Name what broke (one line)",
        "2. Protect life and delivery first",
        "3. Move — do not cancel forever — missed Planning or Sell block",
        "4. Keep Daily Anchors if possible; else resume tomorrow",
    ]
    for s in steps:
        tall_field(ws, row, s, 32, 6)
        row += 1
    row = section(ws, row, "End-of-week Catch-Up Close (up to 60 min only in catch-up/consolidation)", 6)
    end_steps = [
        "1. Run a shortened Close",
        "2. Complete Next-Week Handoff",
        "3. Rebook next Planning Block",
        "4. One sentence: what I will protect first next week",
    ]
    for s in end_steps:
        tall_field(ws, row, s, 32, 6)
        row += 1
    tall_field(ws, row, "This week’s Catch-Up note (when used)", 40, 6)
    footer(ws, row + 2, "Supports Module 4", 6)


def build_tracker(wb: Workbook):
    ws = wb.create_sheet("13-Week Tracker")
    set_widths(ws, {"A": 10, "B": 22, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 28})
    header_block(
        ws,
        "13-Week Tracker",
        "Supports Module 4 · BBOS 13-week journey",
        "Mark each week. Weeks 4, 8, and 12 are consolidation — protect Close; lighter install. "
        "Basic rhythm begins Week 2; pipeline habit deepens Week 3.",
        cols=8,
    )
    # fix date cell for wider sheet
    r = 6
    headers = [
        "Week",
        "Focus",
        "Planning done",
        "Sell block done",
        "Close done",
        "Handoff done",
        "Catch-Up used?",
        "One-line note",
    ]
    style_header_row(ws, r, headers)
    focuses = {
        1: "Foundation / Snapshot",
        2: "Install basic rhythm",
        3: "Pipeline + follow-up habit",
        4: "Consolidation",
        8: "Consolidation",
        12: "Consolidation",
        13: "Close the quarter (Module 8)",
    }
    for w in range(1, 14):
        row = r + w
        label(ws, row, 1, str(w))
        label(ws, row, 2, focuses.get(w, "Operate the rhythm"))
        for c in range(3, 8):
            input_cell(ws, row, c, "")
        input_cell(ws, row, 8, "")
    for col in ("C", "D", "E", "F", "G"):
        dv = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}{r+1}:{col}{r+13}")
    footer(ws, r + 15, "Supports Module 4 (journey)", 8)


def build_example_maya(wb: Workbook):
    ws = wb.create_sheet("Example - Maya")
    set_widths(ws, {"A": 100})
    ws["A1"] = "EXAMPLE — Maya's Property Services (fictional week)"
    ws["A1"].font = font_title
    lines = [
        "",
        "Illustrative only. Not real schedules, clients, addresses, access codes, or business data.",
        "No VelocityMaid, Bornfidis client, or private credentials.",
        "",
        "BUSINESS SNAPSHOT (short)",
        "Business: helps hosts keep properties guest-ready.",
        "First weakness: inconsistent pricing (points to Module 3 — already addressed in journey).",
        "",
        "MY RHYTHM",
        "Planning: Sunday evening ~30 min",
        "Pipeline/Follow-Up: Wednesday late morning ~45 min",
        "Close: Friday afternoon ~35 min",
        "Daily Anchors (3): next-day confirm + access notes with worker; pipeline glance; quality tick",
        "",
        "SUNDAY — PLANNING (~30)",
        "Top 3: deliver booked turnovers; send one priced proposal; protect Wednesday Sell block.",
        "Delivery risk: Thursday morning looks tight.",
        "",
        "MON–TUE — DELIVER + ANCHORS",
        "Turnovers dominate. Quality tick after jobs. Short worker handoffs.",
        "",
        "WEDNESDAY — PIPELINE / FOLLOW-UP (~45)",
        "Update stages; one cadence follow-up; proposal prep (Core rung at Module 3 price).",
        "Should I Say Yes? on new inquiry → Not now (capacity) → Lead + revisit date.",
        "Pricing Checkpoint: no Exceptions; one invoice still to record — noticed, not analyzed.",
        "",
        "WEDNESDAY LATER — DISRUPTION",
        "Urgent same-day turnover. Capacity allows. Sell block paused at ~20 min.",
        "Catch-Up note: finish follow-ups Friday before Close.",
        "",
        "FRIDAY — CATCH-UP + CLOSE (~35 Close)",
        "~20 min Catch-Up follow-ups, then Close inside ceiling.",
        "Next-Week Handoff:",
        "1. Protect Wednesday Sell block + complete proposal follow-up",
        "2. Property manager proposal — cadence day 3",
        "3. Thursday morning load — confirm worker start early",
        "4. Next Planning: Sunday evening ~30 min",
        "",
        "Takeaway: mostly delivery, small protected system blocks, one disruption, recovery without drama.",
    ]
    for i, text in enumerate(lines, start=2):
        cell = ws.cell(i, 1, text)
        if text.isupper() and len(text) < 40:
            cell.font = font_section
        elif text.startswith("Illustrative"):
            cell.font = font_muted
        else:
            cell.font = font_body
    footer(ws, len(lines) + 3, "Example only · Supports Modules 1, 2, 4", 1)


def build_notes(wb: Workbook):
    ws = wb.create_sheet("Notes and Disclaimer")
    set_widths(ws, {"A": 96})
    ws["A1"] = "NOTES / DISCLAIMER"
    ws["A1"].font = font_title
    lines = [
        "",
        "BBOS is a practical operating system for service-business owners.",
        "This workbook is for planning and operating habits only.",
        "Not legal, tax, accounting, HR, medical, investment, or financial advice.",
        "BBOS does not promise revenue, profit, fewer emergencies, or any income result.",
        "",
        "Pricing & Margin Calculator V1.0 remains a separate frozen tool.",
        "This workbook must not duplicate Cost Stack → price formulas.",
        "",
        "Module 4 notices money and pricing signals; Module 5 analyzes them.",
        "Do not use Pricing Checkpoint or Close for runway, cash-flow planning,",
        "or full Core Number Set analysis.",
        "",
        "Pipeline stages and Follow-Up Cadence are defined in Module 2.",
        "Offer structure is defined in Module 2; prices in Module 3.",
        "",
        "Daily Anchors cap: 2–3.",
        "Weekly Close hard ceiling: 45 minutes in a normal week.",
        "Next-Week Handoff is four brief lines — not next week’s Planning Block.",
        "",
        "Privacy: no client passwords, access codes, or confidential third-party data.",
        "Maya's Property Services examples are fictional.",
        "",
        f"Asset: {ASSET}",
        f"{VERSION}",
        "Supports Modules 1, 2, and 4 of the BBOS Main Guide.",
        "Filename: bbos-weekly-operating-rhythm-workbook-v1.xlsx",
    ]
    for i, text in enumerate(lines, start=2):
        ws.cell(i, 1, text).font = font_body


def main():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    build_start_here(wb)
    build_cost_of_disorder(wb)
    build_owner_time_map(wb)
    build_snapshot(wb)
    build_offer_builder(wb)
    build_offer_ladder(wb)
    build_one_pager(wb)
    build_pipeline(wb)
    build_should_i_say_yes(wb)
    build_follow_up_cadence(wb)
    build_my_rhythm(wb)
    build_week_plan(wb)
    build_daily_anchors(wb)
    build_delivery_quality(wb)
    build_pricing_checkpoint(wb)
    build_weekly_close(wb)
    build_catch_up(wb)
    build_tracker(wb)
    build_example_maya(wb)
    build_notes(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print("Sheets:", " | ".join(wb.sheetnames))


if __name__ == "__main__":
    main()
