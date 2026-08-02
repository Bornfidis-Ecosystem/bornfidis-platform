#!/usr/bin/env python3
"""Validate BBOS Weekly Operating Rhythm Workbook V1 structure."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "content" / "bbos" / "bbos-weekly-operating-rhythm-workbook-v1.xlsx"

REQUIRED_SHEETS = [
    "Start Here",
    "Cost of Disorder",
    "Owner Time Map",
    "Business Snapshot",
    "Offer Builder",
    "Offer Ladder",
    "Offer One-Pager",
    "Pipeline Tracker",
    "Should I Say Yes",
    "Follow-Up Cadence",
    "My Rhythm",
    "Week Plan",
    "Daily Anchors",
    "Delivery Quality",
    "Pricing Checkpoint",
    "Weekly Close",
    "Catch-Up Protocol",
    "13-Week Tracker",
    "Example - Maya",
    "Notes and Disclaimer",
]

# User Priority-1 sections → sheet mapping
USER_SECTIONS = {
    "Business Snapshot": "Business Snapshot",
    "Cost of Disorder Audit": "Cost of Disorder",
    "Owner's Time Map": "Owner Time Map",
    "Offer Builder": "Offer Builder",
    "Offer Ladder": "Offer Ladder",
    "Offer One-Pager": "Offer One-Pager",
    "Pipeline Tracker": "Pipeline Tracker",
    "Should I Say Yes?": "Should I Say Yes",
    "Follow-Up Cadence": "Follow-Up Cadence",
    "Weekly Planning Block": "Week Plan",
    "Daily Anchors": "Daily Anchors",
    "Delivery Quality Checklist": "Delivery Quality",
    "Weekly Close": "Weekly Close",
    "Next-Week Handoff": "Weekly Close",  # section inside Close
    "Catch-Up Protocol": "Catch-Up Protocol",
}

FORBIDDEN_PHRASES = [
    "Price = Cost Stack",
    "actual-margin analysis here",  # must warn against, not teach
]

MUST_CONTAIN = {
    "Start Here": ["Pricing & Margin Calculator", "Module 4 notices", "25–40", "2–3"],
    "Pricing Checkpoint": ["Module 4 notices", "Module 5 analyzes", "Pricing & Margin Calculator"],
    "Weekly Close": ["Next-Week Handoff", "H1.", "H4.", "45"],
    "Daily Anchors": ["2–3", "cap"],
    "My Rhythm": ["25–40", "30–60", "25–45", "Rhythm Protection"],
    "Pipeline Tracker": ["Lead", "Qualified", "Proposed", "Won", "Lost"],
    "Notes and Disclaimer": ["Not legal", "does not promise", "Maya"],
    "Example - Maya": ["fictional", "Catch-Up", "Handoff"],
}


def sheet_text(ws) -> str:
    parts = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                parts.append(str(v))
    return "\n".join(parts)


def main() -> int:
    if not XLSX.exists():
        print(f"FAIL: missing {XLSX}")
        return 1

    wb = load_workbook(XLSX, data_only=False)
    errors: list[str] = []

    if wb.sheetnames != REQUIRED_SHEETS:
        errors.append(f"Sheet order/name mismatch.\n  got: {wb.sheetnames}\n  exp: {REQUIRED_SHEETS}")

    for label, sheet in USER_SECTIONS.items():
        if sheet not in wb.sheetnames:
            errors.append(f"Missing sheet for section '{label}' → {sheet}")

    for sheet, needles in MUST_CONTAIN.items():
        text = sheet_text(wb[sheet])
        for n in needles:
            if n.lower() not in text.lower() and n not in text:
                # allow en-dash / hyphen variants
                alt = n.replace("–", "-")
                if alt.lower() not in text.lower() and n not in text:
                    errors.append(f"[{sheet}] missing expected text: {n!r}")

    # Pricing calculator must be referenced, formulas not duplicated as working calc
    start = sheet_text(wb["Start Here"])
    if "Cost Stack ÷" in start or "1 − Target Margin" in start:
        errors.append("Start Here must not embed calculator formulas as a working tool")

    # Confidentiality: Example may name forbidden sources only inside a "not real / no …" disclaimer
    ex = sheet_text(wb["Example - Maya"]).lower()
    if "velocitymaid" in ex and "not real" not in ex and "no velocitymaid" not in ex:
        errors.append("Example sheet appears to use VelocityMaid as real data")
    if "bornfidis.com" in ex:
        errors.append("Example sheet must not include bornfidis.com URLs or live site data")

    # Owner Time Map has SUM formulas
    tm = wb["Owner Time Map"]
    formulas = [tm["B15"].value, tm["C15"].value, tm["D15"].value, tm["E15"].value]
    # row may vary — search for SUM
    found_sum = False
    for row in tm.iter_rows(min_row=1, max_row=30, max_col=5, values_only=True):
        for v in row:
            if isinstance(v, str) and v.startswith("=SUM"):
                found_sum = True
    if not found_sum:
        errors.append("Owner Time Map missing week-total SUM formulas")

    if errors:
        print("VALIDATION FAILED")
        for e in errors:
            print(" -", e)
        return 1

    print("ALL CHECKS PASSED")
    print(f"File: {XLSX}")
    print(f"Sheets ({len(wb.sheetnames)}): {' | '.join(wb.sheetnames)}")
    print("User Priority-1 sections: all mapped")
    print("Pricing Calculator: referenced, not duplicated")
    return 0


if __name__ == "__main__":
    sys.exit(main())
