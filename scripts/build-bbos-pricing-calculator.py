#!/usr/bin/env python3
"""Build BBOS Pricing & Margin Calculator V1 from the frozen Module 3 specification.

Output: content/bbos/bbos-pricing-margin-calculator-v1.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.protection import WorkbookProtection

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "content" / "bbos" / "bbos-pricing-margin-calculator-v1.xlsx"

# Brand tokens (Style Guide)
INK = "1C1C1C"
FOREST = "1A3C34"
GOLD = "C9A84C"
PAPER = "F5F1E6"
WHITE = "FFFFFF"
LINE = "E5E3DC"

thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
gold_border = Border(
    left=Side(style="medium", color=GOLD),
    right=Side(style="medium", color=GOLD),
    top=Side(style="medium", color=GOLD),
    bottom=Side(style="medium", color=GOLD),
)

fill_forest = PatternFill("solid", fgColor=FOREST)
fill_paper = PatternFill("solid", fgColor=PAPER)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_gold_light = PatternFill("solid", fgColor="F8F1D8")

font_title = Font(name="Calibri", size=16, bold=True, color=INK)
font_h = Font(name="Calibri", size=12, bold=True, color=WHITE)
font_label = Font(name="Calibri", size=11, bold=True, color=INK)
font_body = Font(name="Calibri", size=11, color=INK)
font_muted = Font(name="Calibri", size=10, italic=True, color="6B7280")
font_gold = Font(name="Calibri", size=10, bold=True, color=FOREST)

prot_lock = Protection(locked=True)
prot_unlock = Protection(locked=False)


def style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = fill_forest
        cell.font = font_h
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def input_cell(ws, row: int, col: int, value=None, number_format=None):
    cell = ws.cell(row, col, value)
    cell.fill = fill_white
    cell.border = gold_border
    cell.font = font_body
    cell.protection = prot_unlock
    if number_format:
        cell.number_format = number_format
    return cell


def calc_cell(ws, row: int, col: int, value=None, number_format=None):
    cell = ws.cell(row, col, value)
    cell.fill = fill_paper
    cell.border = thin
    cell.font = font_label
    cell.protection = prot_lock
    if number_format:
        cell.number_format = number_format
    return cell


def label(ws, row: int, col: int, text: str):
    cell = ws.cell(row, col, text)
    cell.font = font_label
    cell.protection = prot_lock
    return cell


def note(ws, row: int, col: int, text: str):
    cell = ws.cell(row, col, text)
    cell.font = font_muted
    cell.alignment = Alignment(wrap_text=True)
    cell.protection = prot_lock
    return cell


def build_start_here(wb: Workbook):
    ws = wb.create_sheet("Start Here", 0)
    ws["A1"] = "BBOS Pricing & Margin Calculator — Version 1"
    ws["A1"].font = font_title
    ws.merge_cells("A1:F1")

    lines = [
        "",
        "Purpose",
        "Turn a Cost Stack and Target Margin into a defensible price using the BBOS method:",
        "Price = Cost Stack ÷ (1 − Target Margin)",
        "Margin = (Price − Cost Stack) ÷ Price",
        "",
        "How to use",
        "1. Go to the Inputs sheet. Set your currency symbol first.",
        "2. Enter Core (and optionally Entry / Premium) Cost Stack layers.",
        "3. Enter Target Margin % and Margin Guardrail % (your choices — not universal).",
        "4. Enter Optional item Cost Stacks if needed.",
        "5. Read Results. Exact formula outputs appear first.",
        "6. Optionally enter a practical quoted price (upward rounding only).",
        "7. Open Example to see Maya's Property Services (illustrative USD).",
        "8. Read Notes and Disclaimer before relying on any figure.",
        "",
        "BBOS Pricing Principles",
        "• Price from your costs, not your emotions.",
        "• Protect your standards before protecting the sale.",
        "• Margin is planned — not guessed.",
        "• Discounts are intentional decisions, never automatic reactions.",
        "• A sustainable business serves more people than an exhausted owner.",
        "",
        "Frozen rules (do not change in V1)",
        "• Margin only — never markup.",
        "• Show exact results first; deliberate upward quote rounding allowed.",
        "• Never round down through the Margin Guardrail.",
        "• Optional items never auto-include into Core.",
        "• Contingency is inside the Cost Stack; Target Margin is above it.",
        "",
        "Supports: Module 3 — Pricing With Dignity",
        "For planning only. Not tax, accounting, financial, or legal advice.",
    ]
    for i, text in enumerate(lines, start=2):
        cell = ws.cell(i, 1, text)
        cell.font = font_gold if text.startswith("•") or text in {
            "Purpose", "How to use", "BBOS Pricing Principles", "Frozen rules (do not change in V1)"
        } else font_body
        cell.protection = prot_lock
    ws.column_dimensions["A"].width = 96
    ws.protection.sheet = True
    ws.protection.password = "bbos-v1"


def build_inputs(wb: Workbook):
    ws = wb.create_sheet("Inputs", 1)
    ws["A1"] = "INPUTS — edit gold-bordered cells only"
    ws["A1"].font = font_title
    ws.merge_cells("A1:E1")
    note(ws, 2, 1, "Illustrative defaults follow Maya (USD). Replace with your local inputs.")

    # Currency
    label(ws, 4, 1, "Currency symbol")
    input_cell(ws, 4, 2, "$")
    note(ws, 4, 3, "Currency-neutral: change this symbol for your locale.")

    # Margins
    label(ws, 6, 1, "Target Margin")
    input_cell(ws, 6, 2, 0.30, "0%")
    note(ws, 6, 3, "Planned margin for normal pricing. Your choice — not a universal recommendation.")

    label(ws, 7, 1, "Margin Guardrail")
    input_cell(ws, 7, 2, 0.20, "0%")
    note(ws, 7, 3, "Lowest planned margin without Exception Approval.")

    # Scenario margins
    label(ws, 9, 1, "Scenario A margin")
    input_cell(ws, 9, 2, 0.25, "0%")
    label(ws, 10, 1, "Scenario B margin")
    input_cell(ws, 10, 2, 0.30, "0%")
    label(ws, 11, 1, "Scenario C margin")
    input_cell(ws, 11, 2, 0.35, "0%")

    # Core Cost Stack
    ws["A13"] = "CORE COST STACK"
    style_header_row(ws, 13, 3)
    ws["A14"] = "Layer"
    ws["B14"] = "Amount"
    ws["C14"] = "Notes"
    for col in range(1, 4):
        ws.cell(14, col).font = font_label
        ws.cell(14, col).fill = fill_gold_light

    rows_core = [
        (15, "Direct costs", 20, "Supplies + standard property linen processing"),
        (16, "Labor hours", 3, "Hours to deliver (include your time)"),
        (17, "Labor cost per hour", 20, "What the hour is worth to the business"),
        (18, "Overhead allocation", 15, "Simple per-job share"),
        (19, "Contingency", 10, "Inside Cost Stack — not the same as Target Margin"),
    ]
    for r, name, val, n in rows_core:
        label(ws, r, 1, name)
        fmt = "0.00" if r in (16,) else "0.00"
        if r == 16:
            input_cell(ws, r, 2, val, "0.00")
        elif r == 17:
            input_cell(ws, r, 2, val, "0.00")
        else:
            input_cell(ws, r, 2, val, "0.00")
        note(ws, r, 3, n)

    label(ws, 20, 1, "Labor subtotal (calculated)")
    calc_cell(ws, 20, 2, "=B16*B17", "0.00")

    label(ws, 21, 1, "CORE Cost Stack (calculated)")
    calc_cell(ws, 21, 2, "=B15+B20+B18+B19", "0.00")

    # Entry
    ws["A23"] = "ENTRY COST STACK"
    style_header_row(ws, 23, 3)
    entry = [
        (24, "Direct costs", 12),
        (25, "Labor hours", 2),
        (26, "Labor cost per hour", 20),
        (27, "Overhead allocation", 10),
        (28, "Contingency", 8),
    ]
    for r, name, val in entry:
        label(ws, r, 1, name)
        input_cell(ws, r, 2, val, "0.00")
    label(ws, 29, 1, "Labor subtotal")
    calc_cell(ws, 29, 2, "=B25*B26", "0.00")
    label(ws, 30, 1, "ENTRY Cost Stack")
    calc_cell(ws, 30, 2, "=B24+B29+B27+B28", "0.00")

    # Premium
    ws["A32"] = "PREMIUM COST STACK"
    style_header_row(ws, 32, 3)
    prem = [
        (33, "Direct costs", 28),
        (34, "Labor hours", 6),
        (35, "Labor cost per hour", 20),
        (36, "Overhead allocation", 20),
        (37, "Contingency", 7),
    ]
    for r, name, val in prem:
        label(ws, r, 1, name)
        input_cell(ws, r, 2, val, "0.00")
    label(ws, 38, 1, "Labor subtotal")
    calc_cell(ws, 38, 2, "=B34*B35", "0.00")
    label(ws, 39, 1, "PREMIUM Cost Stack")
    calc_cell(ws, 39, 2, "=B33+B38+B36+B37", "0.00")

    # Optional items
    ws["A41"] = "OPTIONAL ITEMS (each priced separately)"
    style_header_row(ws, 41, 4)
    ws["A42"] = "Optional item name"
    ws["B42"] = "Cost Stack"
    ws["C42"] = "Notes"
    for col in range(1, 4):
        ws.cell(42, col).font = font_label
        ws.cell(42, col).fill = fill_gold_light

    optionals = [
        (43, "Inside-oven / fridge deep clean", 21, "Labor-heavy add-on"),
        (44, "Additional guest/owner laundry", 14, "Outside normal turnover linen processing"),
        (45, "Consumables restock", 7, ""),
        (46, "", None, "Add more rows as needed"),
        (47, "", None, ""),
    ]
    for r, name, val, n in optionals:
        input_cell(ws, r, 1, name)
        if val is None:
            input_cell(ws, r, 2, None, "0.00")
        else:
            input_cell(ws, r, 2, val, "0.00")
        note(ws, r, 3, n)

    # Practical quote override
    ws["A49"] = "PRACTICAL QUOTED PRICE (optional upward round)"
    style_header_row(ws, 49, 3)
    label(ws, 50, 1, "Practical Core quote (optional)")
    input_cell(ws, 50, 2, None, "0.00")
    note(ws, 50, 3, "Leave blank to use exact Core price. If filled, must be >= exact Guardrail floor.")

    label(ws, 51, 1, "Exception flag (Yes/No)")
    input_cell(ws, 51, 2, "No")
    note(ws, 51, 3, "Set Yes only with documented Exception Approval if quoting below Guardrail.")

    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B51"])

    for col, width in {"A": 42, "B": 18, "C": 55, "D": 20}.items():
        ws.column_dimensions[col].width = width

    # Named-style protection: unlock inputs already set; protect sheet
    ws.protection.sheet = True
    ws.protection.password = "bbos-v1"
    ws.protection.enable()


def build_results(wb: Workbook):
    ws = wb.create_sheet("Results", 2)
    ws["A1"] = "RESULTS — calculated (do not edit)"
    ws["A1"].font = font_title
    ws.merge_cells("A1:D1")
    note(ws, 2, 1, 'All exact prices use: Cost ÷ (1 − margin). Status never relies on color alone.')

    ws["A4"] = "CORE"
    style_header_row(ws, 4, 3)

    status_formula = (
        '=IF(OR(NOT(ISNUMBER(B6)),NOT(ISNUMBER(B7)),B7>=1,B8>=1),"Check inputs",'
        'IF(AND(ISNUMBER(Inputs!B50),Inputs!B50<B11,Inputs!B51<>"Yes"),"Below guardrail",'
        'IF(ABS(IF(ISNUMBER(B13),B13,B10)-B7)<=0.005,"At target",'
        'IF(IF(ISNUMBER(B13),B13,B10)<B8,"Below guardrail",'
        'IF(IF(ISNUMBER(B13),B13,B10)<B7,"Between guardrail and target","At or above target")))))'
    )
    rows = [
        (5, "Currency", "=Inputs!B4", None),
        (6, "Core Cost Stack", "=Inputs!B21", "0.00"),
        (7, "Target Margin", "=Inputs!B6", "0.00%"),
        (8, "Margin Guardrail", "=Inputs!B7", "0.00%"),
        (9, "Exact Core price (Target)", '=IF(Inputs!B6>=1,"—",Inputs!B21/(1-Inputs!B6))', "0.00"),
        (10, "Verified margin at exact price", '=IF(OR(B9="—",NOT(ISNUMBER(B9))),"—",(B9-B6)/B9)', "0.00%"),
        (11, "Exact Guardrail floor", '=IF(Inputs!B7>=1,"—",Inputs!B21/(1-Inputs!B7))', "0.00"),
        (12, "Practical Core quote (from Inputs)", '=IF(Inputs!B50="",B9,Inputs!B50)', "0.00"),
        (13, "Verified margin at practical quote", '=IF(OR(B12="",B12="—",NOT(ISNUMBER(B12))),"—",(B12-B6)/B12)', "0.00%"),
        (14, "Status", status_formula, None),
    ]

    for r, name, formula, fmt in rows:
        label(ws, r, 1, name)
        calc_cell(ws, r, 2, formula, fmt)

    note(
        ws,
        15,
        1,
        "Rounding rule: exact results first. You may round a quote UP. Never round down through the Guardrail.",
    )

    # Scenarios
    ws["A17"] = "SCENARIO COMPARISON (Core Cost Stack)"
    style_header_row(ws, 17, 4)
    ws["A18"] = "Scenario"
    ws["B18"] = "Margin"
    ws["C18"] = "Exact price"
    ws["D18"] = "Notes"
    for c in range(1, 5):
        ws.cell(18, c).font = font_label
        ws.cell(18, c).fill = fill_gold_light

    label(ws, 19, 1, "A")
    calc_cell(ws, 19, 2, "=Inputs!B9", "0%")
    calc_cell(ws, 19, 3, "=IF(B19>=1,\"—\",Inputs!B21/(1-B19))", "0.00")
    note(ws, 19, 4, "Tighter than target (example)")

    label(ws, 20, 1, "B")
    calc_cell(ws, 20, 2, "=Inputs!B10", "0%")
    calc_cell(ws, 20, 3, "=IF(B20>=1,\"—\",Inputs!B21/(1-B20))", "0.00")
    note(ws, 20, 4, "Usually matches Target Margin")

    label(ws, 21, 1, "C")
    calc_cell(ws, 21, 2, "=Inputs!B11", "0%")
    calc_cell(ws, 21, 3, "=IF(B21>=1,\"—\",Inputs!B21/(1-B21))", "0.00")
    note(ws, 21, 4, "Maya example: exact $161.54; intentional quote $162")

    # Ladder
    ws["A23"] = "OFFER LADDER PRICES (at Target Margin)"
    style_header_row(ws, 23, 4)
    ws["A24"] = "Rung"
    ws["B24"] = "Cost Stack"
    ws["C24"] = "Exact price"
    ws["D24"] = "Verified margin"
    for c in range(1, 5):
        ws.cell(24, c).font = font_label
        ws.cell(24, c).fill = fill_gold_light

    label(ws, 25, 1, "Entry")
    calc_cell(ws, 25, 2, "=Inputs!B30", "0.00")
    calc_cell(ws, 25, 3, "=IF(Inputs!B6>=1,\"—\",B25/(1-Inputs!B6))", "0.00")
    calc_cell(ws, 25, 4, '=IF(C25="—","—",(C25-B25)/C25)', "0.00%")

    label(ws, 26, 1, "Core")
    calc_cell(ws, 26, 2, "=Inputs!B21", "0.00")
    calc_cell(ws, 26, 3, "=B9", "0.00")
    calc_cell(ws, 26, 4, "=B10", "0.00%")

    label(ws, 27, 1, "Premium")
    calc_cell(ws, 27, 2, "=Inputs!B39", "0.00")
    calc_cell(ws, 27, 3, "=IF(Inputs!B6>=1,\"—\",B27/(1-Inputs!B6))", "0.00")
    calc_cell(ws, 27, 4, '=IF(C27="—","—",(C27-B27)/C27)', "0.00%")

    # Optionals
    ws["A29"] = "OPTIONAL ITEM PRICES (at Target Margin)"
    style_header_row(ws, 29, 4)
    ws["A30"] = "Optional item"
    ws["B30"] = "Cost Stack"
    ws["C30"] = "Exact price"
    ws["D30"] = "Verified margin"
    for c in range(1, 5):
        ws.cell(30, c).font = font_label
        ws.cell(30, c).fill = fill_gold_light

    for i, r in enumerate(range(31, 36)):
        src = 43 + i
        calc_cell(ws, r, 1, f'=Inputs!A{src}')
        calc_cell(ws, r, 2, f'=IF(Inputs!B{src}="","" ,Inputs!B{src})', "0.00")
        calc_cell(
            ws,
            r,
            3,
            f'=IF(OR(B{r}="",Inputs!B6>=1),"",B{r}/(1-Inputs!B6))',
            "0.00",
        )
        calc_cell(
            ws,
            r,
            4,
            f'=IF(OR(C{r}="",C{r}=0),"",(C{r}-B{r})/C{r})',
            "0.00%",
        )

    note(
        ws,
        37,
        1,
        "Optional items never auto-include into Core. Price each separately. Not Included work stays out of the Cost Stack.",
    )

    for col, width in {"A": 40, "B": 22, "C": 18, "D": 40}.items():
        ws.column_dimensions[col].width = width

    ws.protection.sheet = True
    ws.protection.password = "bbos-v1"
    ws.protection.enable()


def build_example(wb: Workbook):
    ws = wb.create_sheet("Example", 3)
    ws["A1"] = "EXAMPLE — Maya's Property Services (illustrative USD)"
    ws["A1"].font = font_title
    ws.merge_cells("A1:D1")
    note(
        ws,
        2,
        1,
        "Illustrative example in USD. Replace the currency symbol and figures with your own local inputs. "
        "Figures are fictional — not recommended rates. Matches Module 3 Main Guide exactly.",
    )

    data = [
        "",
        "CORE COST STACK",
        "Direct costs | $20 | Supplies + standard property linen processing",
        "Labor | 3 hours × $20/hour = $60 | Owner time counts",
        "Overhead allocation | $15 | Simple per-job share",
        "Contingency | $10 | Inside Cost Stack",
        "CORE Cost Stack | $105",
        "",
        "MARGINS",
        "Target Margin | 30%",
        "Margin Guardrail | 20%",
        "",
        "CORE PRICE",
        "Exact Core price | $105 ÷ (1 − 0.30) = $150.00",
        "Verified margin | ($150 − $105) ÷ $150 = 30%",
        "Exact Guardrail floor | $105 ÷ (1 − 0.20) = $131.25",
        "Practical minimum quote | $132 (rounded upward so she does not move below the Guardrail)",
        "",
        "SCENARIOS (Cost Stack $105)",
        "A | 25% | Exact $140.00",
        "B | 30% | Exact $150.00 (chosen)",
        "C | 35% | Exact $161.54 → intentional quote $162 (upward round; post-round margin ≈ 35.2%)",
        "",
        "OFFER LADDER (Target Margin 30%)",
        "Entry Cost Stack $70 → Exact price $100",
        "Core Cost Stack $105 → Exact price $150",
        "Premium Cost Stack $175 → Exact price $250",
        "",
        "OPTIONALS (Target Margin 30%)",
        "Oven/fridge Cost Stack $21 → $30",
        "Additional guest/owner laundry Cost Stack $14 → $20",
        "Consumables restock Cost Stack $7 → $10",
        "",
        "DISCOUNT / EXCEPTION (process illustration only)",
        "Ask $120 → below exact Guardrail floor $131.25 → decline",
        "Bounded exception $140 for first 4 Core turnovers (margin 25%) with end date, then return to $150",
    ]
    for i, text in enumerate(data, start=3):
        cell = ws.cell(i, 1, text)
        cell.font = font_label if text.isupper() or text in {
            "CORE COST STACK", "MARGINS", "CORE PRICE", "SCENARIOS (Cost Stack $105)",
            "OFFER LADDER (Target Margin 30%)", "OPTIONALS (Target Margin 30%)",
            "DISCOUNT / EXCEPTION (process illustration only)",
        } or (text and text == text.upper() and len(text) < 40) else font_body
        if text in {
            "CORE COST STACK",
            "MARGINS",
            "CORE PRICE",
            "SCENARIOS (Cost Stack $105)",
            "OFFER LADDER (Target Margin 30%)",
            "OPTIONALS (Target Margin 30%)",
            "DISCOUNT / EXCEPTION (process illustration only)",
        }:
            cell.font = font_gold
        cell.protection = prot_lock
    ws.column_dimensions["A"].width = 110
    ws.protection.sheet = True
    ws.protection.password = "bbos-v1"


def build_notes(wb: Workbook):
    ws = wb.create_sheet("Notes and Disclaimer", 4)
    ws["A1"] = "NOTES / DISCLAIMER"
    ws["A1"].font = font_title
    lines = [
        "",
        "For planning only. Not tax, accounting, financial, investment, or legal advice.",
        "BBOS does not guarantee profit, revenue, or any specific business result.",
        "",
        "Margin is not markup.",
        "Margin = (Price − Cost) ÷ Price",
        "Markup = (Price − Cost) ÷ Cost",
        "This calculator uses margin only.",
        "",
        "Contingency protects the job from unexpected costs.",
        "Target Margin rewards the business after all planned costs have been covered.",
        "",
        "No client data, real Bornfidis pricing, VelocityMaid data, or private credentials belong in this file.",
        "Maya's Property Services figures are fictional and illustrative.",
        "",
        "Competitor prices are not your Cost Stack. Use this calculator first;",
        "competitor awareness is a final reality check only.",
        "",
        "Supports Module 3 — Pricing With Dignity | BBOS Version 1",
        "The Main Guide defines this spreadsheet. The spreadsheet must not invent pricing behavior.",
    ]
    for i, text in enumerate(lines, start=2):
        cell = ws.cell(i, 1, text)
        cell.font = font_body
        cell.protection = prot_lock
    ws.column_dimensions["A"].width = 96
    ws.protection.sheet = True
    ws.protection.password = "bbos-v1"


def main():
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    build_start_here(wb)
    build_inputs(wb)
    build_results(wb)
    build_example(wb)
    build_notes(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
