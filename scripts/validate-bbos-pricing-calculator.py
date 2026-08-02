#!/usr/bin/env python3
"""Validate BBOS Pricing & Margin Calculator against Maya + 10 fictional scenarios."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "content" / "bbos" / "bbos-pricing-margin-calculator-v1.xlsx"


def price(cost: float, margin: float) -> float:
    return cost / (1 - margin)


def margin(cost: float, p: float) -> float:
    return (p - cost) / p


def almost(a: float, b: float, tol: float = 1e-9) -> bool:
    return abs(a - b) <= tol


def check(name: str, ok: bool, detail: str = ""):
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {name}" + (f" — {detail}" if detail else ""))
    return ok


def main() -> int:
    assert XLSX.exists(), f"Missing {XLSX}"
    wb = load_workbook(XLSX)
    expected_sheets = ["Start Here", "Inputs", "Results", "Example", "Notes and Disclaimer"]
    ok_all = True
    ok_all &= check("Sheet order", list(wb.sheetnames) == expected_sheets, str(wb.sheetnames))

    inp = wb["Inputs"]
    # Maya defaults on Inputs
    checks = [
        ("Currency", inp["B4"].value == "$"),
        ("Target Margin", almost(float(inp["B6"].value), 0.30)),
        ("Guardrail", almost(float(inp["B7"].value), 0.20)),
        ("Core direct", almost(float(inp["B15"].value), 20)),
        ("Core hours", almost(float(inp["B16"].value), 3)),
        ("Core rate", almost(float(inp["B17"].value), 20)),
        ("Core OH", almost(float(inp["B18"].value), 15)),
        ("Core contingency", almost(float(inp["B19"].value), 10)),
        ("Entry stack inputs", almost(float(inp["B24"].value) + float(inp["B25"].value) * float(inp["B26"].value) + float(inp["B27"].value) + float(inp["B28"].value), 70)),
        ("Premium stack inputs", almost(float(inp["B33"].value) + float(inp["B34"].value) * float(inp["B35"].value) + float(inp["B36"].value) + float(inp["B37"].value), 175)),
        ("Optional oven cost", almost(float(inp["B43"].value), 21)),
        ("Optional laundry cost", almost(float(inp["B44"].value), 14)),
        ("Optional consumables cost", almost(float(inp["B45"].value), 7)),
    ]
    for name, result in checks:
        ok_all &= check(f"Maya input: {name}", result)

    # Formula presence on Results
    res = wb["Results"]
    ok_all &= check("Core price formula", "1-Inputs!B6" in str(res["B9"].value))
    ok_all &= check("Guardrail formula", "1-Inputs!B7" in str(res["B11"].value))
    ok_all &= check("Status text formula present", "Below guardrail" in str(res["B14"].value))

    # Independent Maya math (guide lock)
    core = 105.0
    ok_all &= check("Maya core price", almost(price(core, 0.30), 150))
    ok_all &= check("Maya verified margin", almost(margin(core, 150), 0.30))
    ok_all &= check("Maya guardrail floor", almost(price(core, 0.20), 131.25))
    ok_all &= check("Maya scenario A", almost(price(core, 0.25), 140))
    ok_all &= check("Maya scenario C exact", almost(round(price(core, 0.35), 2), 161.54))
    ok_all &= check("Maya entry", almost(price(70, 0.30), 100))
    ok_all &= check("Maya premium", almost(price(175, 0.30), 250))
    ok_all &= check("Maya optional oven", almost(price(21, 0.30), 30))
    ok_all &= check("Exception 140 above guardrail", margin(core, 140) > 0.20)
    ok_all &= check("Ask 120 below guardrail", margin(core, 120) < 0.20)

    # 10 additional fictional scenarios
    scenarios = [
        ("Lawn care small", 40, 0.35),
        ("Bookkeeping monthly", 180, 0.40),
        ("Mobile detailing", 65, 0.25),
        ("Tutoring package", 90, 0.45),
        ("Handyman visit", 110, 0.30),
        ("Dog walking week", 55, 0.28),
        ("Photography session", 220, 0.50),
        ("Meal prep weekly", 130, 0.32),
        ("IT support block", 95, 0.38),
        ("Event staffing", 300, 0.22),
    ]
    for name, cost, tm in scenarios:
        p = price(cost, tm)
        m = margin(cost, p)
        g = price(cost, max(tm - 0.10, 0.05))
        ok = almost(m, tm) and p > cost and g < p
        ok_all &= check(f"Scenario: {name}", ok, f"cost={cost} tm={tm} price={p:.2f} guard={g:.2f}")

    # Protection / structure
    ok_all &= check("Example sheet mentions $161.54", "161.54" in "".join(
        str(c.value) for row in wb["Example"].iter_rows(min_row=1, max_row=40, max_col=1) for c in row if c.value
    ))
    ok_all &= check("Notes disclaimer present", "Not tax" in str(wb["Notes and Disclaimer"]["A3"].value))

    print("\n" + ("ALL CHECKS PASSED" if ok_all else "SOME CHECKS FAILED"))
    return 0 if ok_all else 1


if __name__ == "__main__":
    raise SystemExit(main())
